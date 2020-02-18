import os, json
import sys
import pandas as pd
import openpyxl
from pandas import ExcelWriter
from pandas.io.json import json_normalize
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl import Workbook
import simplejson


class PrettyFloat(float):
    def __repr__(self):
        return '%.15g' % self

def pretty_floats(obj):
    if isinstance(obj, float):
        return PrettyFloat(obj)
    elif isinstance(obj, dict):
        return dict((k, pretty_floats(v)) for k, v in obj.items())
    elif isinstance(obj, (list, tuple)):
        return list(map(pretty_floats, obj)) # in Python3 do: list(map(pretty_floats, obj))
    return obj


#file = sys.argv[2]
#path = sys.argv[1]
wb = openpyxl.Workbook()
wb.save(sys.argv[2])

# this finds our json files
#path_to_json = 'E:/Projekt/Jsons/Skeletor/andere/'
path_to_json = sys.argv[1]
json_files = [pos_json for pos_json in os.listdir(path_to_json) if pos_json.endswith('.json')]


# here I define my pandas Dataframe with the columns I want to get from the json
#jsons_data = pd.DataFrame(columns=['pose','handl','handr'])
pose_data = pd.DataFrame(columns=['pose'])
#handl_data = pd.DataFrame(columns=['handl'])
#handr_data = pd.DataFrame(columns=['handr'])

# we need both the json and an index number so use enumerate()
for index, js in enumerate(json_files):
    with open(os.path.join(path_to_json, js)) as json_file:
        json_text = json.load(json_file)

        
        # here you need to know the layout of your json and each json has to have
        # the same structure (obviously not the structure I have here)
        pose = json_text[0]['keypoints']
        #handl = json_text['people'][0]['hand_left_keypoints_2d']
        #handr = json_text['people'][0]['hand_right_keypoints_2d']

            #print (len(pose))
            
            # here I push a list of data into a pandas DataFrame at row given by 'index'
            #jsons_data.loc[index] = [pose, handl, handr]
        pose_data = pd.DataFrame(pose, columns=['pose'])
        #handl_data = pd.DataFrame(pose, columns=['handl'])
        #handr_data = pd.DataFrame(pose, columns=['handr'])
        
            #pose_data.loc[index] = [pose]
            #handl_data.loc[index] = [handl]
            #handr_data.loc[index] = [handr]
         
        
        #print(pose_data)
        #print(handl_data)
        #print(handr_data)
        
        
        book = load_workbook(sys.argv[2])
        book.alignment = Alignment(horizontal="right")
        writer = pd.ExcelWriter(sys.argv[2], engine='openpyxl') 
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        pose_data.to_excel(writer, 'Main', startrow = 0, startcol = index, 
                             header = True, index = False)
        #handl_data.to_excel(writer, 'Main', startrow = 76, startcol = index, 
                             #header = True, index = False)
        #handr_data.to_excel(writer, 'Main', startrow = 152, startcol = index, 
                            #header = True, index = False)
        # Formating
        #value_fmt = book.number_format({'num_format': '#,###0.000'})
        #book.number_format = '0.000E+000'    
        book._named_styles['Normal'].number_format = '#,###0.000'
    writer.save()

    
    



